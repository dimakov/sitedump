from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter
import json
import sys
import os
import time



def main(supplied_f):

    f_name = ''
    if len(supplied_f) == 1:
        f_name = 'yad2.data'
    else:
        f_name = supplied_f[1]
    try:
        f = open(f_name, 'r');
    except IOError as e:
        print "I/O error({0}): {1}".format(e.errno, e.strerror)
    except:
        print "Unexpected error:", sys.exc_info()[0]
        raise

    # load json data from opened file
    data = json.load(f)
    try:
        wb = load_workbook('yad2-spark.xlsx')
    except IOError:
        print "File not created, creating file with name %s" % f_name
        wb = Workbook()
        wb.save('yad2-spark.xlsx')
        wb = load_workbook('yad2-spark.xlsx')

    # initialize file
    ws = wb.create_sheet()
    ws.title = str(time.strftime("%d-%m-%Y_%Hh%Mm%Ss"))

    #header initialize
    ws.cell(row=1, column=1).value = "Car name"
    ws.cell(row=1, column=2).value = "Hand"
    ws.cell(row=1, column=3).value = "Volume"
    ws.cell(row=1, column=4).value = "Year"
    ws.cell(row=1, column=5).value = "Price"
    ws.cell(row=1, column=6).value = "Min Price"
    ws.cell(row=1, column=7).value = "Max Price"
    ws.cell(row=1, column=8).value = "URL"
    ws.column_dimensions[get_column_letter(1)].width = 16
    ws.column_dimensions[get_column_letter(8)].width = 83
    ws.column_dimensions[get_column_letter(2)].style.alignment.horizontal = 'center'

    for i, item in enumerate(data):
        ws.cell(row = i + 2, column=1).value = data[item]['car_name'].encode('utf-8')
        ws.cell(row = i + 2, column=2).value = data[item]['hand']
        ws.cell(row = i + 2, column=3).value = data[item]['volume']
        ws.cell(row = i + 2, column=4).value = data[item]['year']
        ws.cell(row = i + 2, column=5).value = data[item]['price']
        ws.cell(row = i + 2, column=6).value = data[item]['min_price']
        ws.cell(row = i + 2, column=7).value = data[item]['max_price']
        ws.cell(row = i + 2, column=8).value = data[item]['url']

    wb.save('yad2-spark.xlsx')
    f.close()

if __name__=="__main__":
    main(sys.argv)